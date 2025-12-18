import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

class GameDataSaver:
    def __init__(self, filename="game_data.xlsx"):
        self.filename = filename

        folder = os.path.dirname(self.filename)
        if folder:
            os.makedirs(folder, exist_ok=True)

    def _get_sheet(self, sheet_name, header):
        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws.append(header)
            wb.save(self.filename)

        wb = load_workbook(self.filename)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(header)
        else:
            ws = wb[sheet_name]

        return wb, ws

    def save_multiplayer(self, player1, player2, level):
        sheet_name = "Multiplayer"
        header = ["Tanggal", "Player1", "Player2", "Level", "Pemenang", "HighScore"]

        wb, ws = self._get_sheet(sheet_name, header)

        if player1 > player2:
            winner = "Player1"
        elif player2 > player1:
            winner = "Player2"
        else:
            winner = "Draw"

        highscore = max(player1, player2)

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            player1,
            player2,
            level,
            winner,
            highscore
        ])

        rows = list(ws.iter_rows(values_only=True))[1:]
        rows.sort(key=lambda r: r[5], reverse=True)

        ws.delete_rows(2, ws.max_row)
        for r in rows:
            ws.append(r)

        wb.save(self.filename)

    def save_solo(self, score, missed, level):
        sheet_name = "Soloplayer"
        header = ["Tanggal", "Score", "Missed", "Level"]

        wb, ws = self._get_sheet(sheet_name, header)

        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            score,
            missed,
            level
        ])

        rows = list(ws.iter_rows(values_only=True))[1:]
        rows.sort(key=lambda r: r[1], reverse=True)

        ws.delete_rows(2, ws.max_row)
        for r in rows:
            ws.append(r)

        wb.save(self.filename)
